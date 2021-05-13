from openpyxl import load_workbook

wb = load_workbook('tabla.xlsx')
#sheet  = wb['Hoja1']

#Método alternativo de captura de hoja
sheets = wb.sheetnames
sheet = wb[sheets[0]]

#Capturar una columna entera
columna = sheet['A']
for valor in columna:
   print(valor.value)
   
#Capturo un fila entera
fila = sheet[2]
for valor in fila:
   print(valor.value)
   
#Recorrer todas las filas
filas = sheet.rows
for fila in filas:
   print("Fila:")
   for celda in fila:
      print(celda.value)
      
#Recorrer todas las columnas
columnas = sheet.columns
for columna in columnas:
   print("Columna:")
   for celda in columna:
      print(celda.value)

#Recorrer todos los valores
for y in range(1, sheet.max_row+1):
   for x in range(1, sheet.max_column+1):
      print(sheet.cell(y,x).value)
   
###Transformar a csv
with open('datos_excel.csv', 'w') as csv:
   for y in range(1, sheet.max_row+1):
      for x in range(1, sheet.max_column+1):
         if x == sheet.max_column:
            csv.write(str(sheet.cell(y,x).value) + '\n')
         else:
            csv.write(str(sheet.cell(y,x).value) + ',')






